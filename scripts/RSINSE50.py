"""
NSE Stock Screener - Nifty 50
==============================
Filters stocks based on:
- Monthly RSI > 50
- Weekly RSI  > 60
- Daily RSI   > 45
- Current price > 200 EMA for last 4 weeks (20 trading days)
- 150 EMA > 200 EMA
- Daily ADX   > 20
- Avg Daily Volume > 100,000
- Free Float Market Cap > 3000 Crore
"""

import os, sys
BASE_DIR    = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INPUT_CSV   = os.path.join(BASE_DIR, "stock_lists", "nifty50list.csv")
OUTPUT_XLSX = os.path.join(BASE_DIR, "reports",     "Nifty50_stocks.xlsx")
OUTPUT_JSON = os.path.join(BASE_DIR, "reports",     "Nifty50_data.json")

import pandas as pd
import numpy as np
import yfinance as yf
import time
import json
import warnings
warnings.filterwarnings("ignore")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── helpers ──────────────────────────────────────────────────────────────────

class _SuppressOutput:
    def __enter__(self):
        import io
        self._orig_stdout = sys.stdout
        self._orig_stderr = sys.stderr
        sys.stdout = io.StringIO()
        sys.stderr = io.StringIO()
        return self
    def __exit__(self, *_):
        sys.stdout = self._orig_stdout
        sys.stderr = self._orig_stderr


def safe_download(symbol, period, interval, retries=3, pause=2.0):
    for attempt in range(1, retries + 1):
        try:
            with _SuppressOutput():
                df = yf.download(symbol, period=period, interval=interval,
                                 auto_adjust=True, progress=False, threads=False)
            if df is None or df.empty:
                raise ValueError("empty result")
            if isinstance(df.columns, pd.MultiIndex):
                df.columns = df.columns.get_level_values(0)
            df = df.dropna(axis=1, how="all")
            return df
        except Exception:
            if attempt < retries:
                time.sleep(pause * attempt)
    return pd.DataFrame()


def compute_rsi(series, period=14):
    delta    = series.diff()
    gain     = delta.clip(lower=0)
    loss     = -delta.clip(upper=0)
    avg_gain = gain.ewm(com=period - 1, min_periods=period).mean()
    avg_loss = loss.ewm(com=period - 1, min_periods=period).mean()
    rs       = avg_gain / avg_loss.replace(0, np.nan)
    rsi      = 100 - (100 / (1 + rs))
    return round(float(rsi.iloc[-1]), 2)


def compute_ema(series, period):
    return series.ewm(span=period, adjust=False).mean()


def compute_adx(high, low, close, period=14):
    tr = pd.concat([
        high - low,
        (high - close.shift()).abs(),
        (low  - close.shift()).abs()
    ], axis=1).max(axis=1)
    dm_pos = high.diff().clip(lower=0).where(high.diff() > low.diff().abs(), 0)
    dm_neg = low.diff().abs().clip(lower=0).where(low.diff().abs() > high.diff(), 0)
    atr    = tr.ewm(span=period, adjust=False).mean()
    di_pos = 100 * dm_pos.ewm(span=period, adjust=False).mean() / atr
    di_neg = 100 * dm_neg.ewm(span=period, adjust=False).mean() / atr
    dx  = (100 * (di_pos - di_neg).abs() / (di_pos + di_neg)).replace([np.inf, -np.inf], np.nan)
    adx = dx.ewm(span=period, adjust=False).mean()
    return round(float(adx.iloc[-1]), 2)


def price_above_200ema_last_4weeks(close, ema200, trading_days=20):
    if len(close) < trading_days:
        return False
    return bool((close.iloc[-trading_days:] > ema200.iloc[-trading_days:]).all())


def load_symbol_list(csv_path):
    df = pd.read_csv(csv_path)
    for col in ["Symbol", "SYMBOL", "symbol", "Ticker", "ticker", "NSE Code"]:
        if col in df.columns:
            symbols = df[col].dropna().str.strip().tolist()
            return [s + ".NS" if not s.endswith(".NS") else s for s in symbols]
    raise ValueError(f"No symbol column found in {csv_path}. Columns: {list(df.columns)}")


def fetch_and_process(symbol):
    ticker = symbol.replace(".NS", "")
    rsi_row = screen_row = None
    try:
        daily   = safe_download(symbol, "2y",  "1d")
        weekly  = safe_download(symbol, "5y",  "1wk")
        monthly = safe_download(symbol, "10y", "1mo")

        if not daily.empty and len(daily) >= 14:
            close_d       = daily["Close"].dropna()
            current_price = float(close_d.iloc[-1])
            rsi_daily   = compute_rsi(close_d) if len(close_d) >= 14 else None
            rsi_weekly  = compute_rsi(weekly["Close"].dropna())  if (not weekly.empty  and len(weekly)  >= 14) else None
            rsi_monthly = compute_rsi(monthly["Close"].dropna()) if (not monthly.empty and len(monthly) >= 14) else None
            rsi_row = {
                "Symbol":        ticker,
                "Current_Price": round(current_price, 2),
                "Daily_RSI":     rsi_daily,
                "Weekly_RSI":    rsi_weekly,
                "Monthly_RSI":   rsi_monthly,
            }

        if daily.empty or len(daily) < 210:
            return rsi_row, None

        close_d = daily["Close"].dropna()
        high_d  = daily["High"].dropna()
        low_d   = daily["Low"].dropna()
        vol_d   = daily["Volume"].dropna()

        ema150 = compute_ema(close_d, 150)
        ema200 = compute_ema(close_d, 200)

        current_price = float(close_d.iloc[-1])
        ema150_val    = float(ema150.iloc[-1])
        ema200_val    = float(ema200.iloc[-1])

        rsi_daily   = rsi_row["Daily_RSI"]   if rsi_row else compute_rsi(close_d)
        if weekly.empty or len(weekly) < 20:  return rsi_row, None
        rsi_weekly  = rsi_row["Weekly_RSI"]  if rsi_row else compute_rsi(weekly["Close"].dropna())
        if monthly.empty or len(monthly) < 20: return rsi_row, None
        rsi_monthly = rsi_row["Monthly_RSI"] if rsi_row else compute_rsi(monthly["Close"].dropna())

        adx        = compute_adx(high_d, low_d, close_d)
        avg_volume = float(vol_d.tail(30).mean())
        info       = yf.Ticker(symbol).info
        market_cap    = info.get("marketCap", 0) or 0
        market_cap_cr = market_cap / 1e7

        if rsi_monthly is None or rsi_monthly <= 50: return rsi_row, None
        if rsi_weekly  is None or rsi_weekly  <= 60: return rsi_row, None
        if rsi_daily   is None or rsi_daily   <= 45: return rsi_row, None
        if not price_above_200ema_last_4weeks(close_d, ema200): return rsi_row, None
        if ema150_val  <= ema200_val:   return rsi_row, None
        if adx         <= 20:           return rsi_row, None
        if avg_volume  <= 100_000:      return rsi_row, None
        if market_cap_cr <= 3000:       return rsi_row, None

        screen_row = {
            "Symbol":        ticker,
            "Current_Price": round(current_price, 2),
            "Daily_RSI":     rsi_daily,
            "Weekly_RSI":    rsi_weekly,
            "Monthly_RSI":   rsi_monthly,
            "ADX":           adx,
            "EMA_150":       round(ema150_val, 2),
            "EMA_200":       round(ema200_val, 2),
            "Avg_Volume":    int(avg_volume),
            "MarketCap_Cr":  round(market_cap_cr, 0),
        }
    except Exception as e:
        print(f"  [ERROR] {symbol}: {e}")
    return rsi_row, screen_row


HEADER_FILL  = PatternFill("solid", start_color="1F4E79")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
DATA_FONT    = Font(name="Arial", size=10)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN   = Alignment(horizontal="left",   vertical="center")
LIGHT_GREEN  = PatternFill("solid", start_color="C6EFCE")
_thin        = Side(style="thin", color="BFBFBF")
THIN_BORDER  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _style_header_row(ws, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN; cell.border = THIN_BORDER


def _style_data_row(ws, row, col_count, fill=None):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = DATA_FONT
        cell.alignment = CENTER_ALIGN if col > 1 else LEFT_ALIGN
        cell.border = THIN_BORDER
        if fill: cell.fill = fill


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_excel(results, rsi_all, output_path):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    df_screened = pd.DataFrame(results).sort_values("Daily_RSI", ascending=False) if results else pd.DataFrame()
    df_rsi      = pd.DataFrame(rsi_all).sort_values("Daily_RSI", ascending=False, na_position="last")
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Screener Results"
    ws1.freeze_panes = "A2"
    full_cols = ["Symbol","Chart","Current_Price","Daily_RSI","Weekly_RSI",
                 "Monthly_RSI","ADX","EMA_150","EMA_200","Avg_Volume","MarketCap_Cr"]
    ws1.append(full_cols)
    _style_header_row(ws1, len(full_cols))

    if not df_screened.empty:
        for _, row in df_screened.iterrows():
            ws1.append([row[c] for c in ["Symbol","Current_Price","Daily_RSI","Weekly_RSI",
                                          "Monthly_RSI","ADX","EMA_150","EMA_200","Avg_Volume","MarketCap_Cr"]])
            data_row = ws1.max_row
            for col in range(11, 1, -1):
                ws1.cell(row=data_row, column=col+1).value = ws1.cell(row=data_row, column=col).value
            tv_url = f"https://www.tradingview.com/chart/?symbol=NSE:{row['Symbol']}"
            chart_cell = ws1.cell(row=data_row, column=2)
            chart_cell.hyperlink = tv_url
            chart_cell.value = "Chart"
            chart_cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")
            chart_cell.alignment = CENTER_ALIGN
            _style_data_row(ws1, data_row, len(full_cols))
            chart_cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")

    _set_col_widths(ws1, [16,9,14,11,13,14,10,12,12,14,14])

    ws2 = wb.create_sheet(title="RSI Summary")
    ws2.freeze_panes = "A2"
    summary_cols = ["Ticker","Chart","Current_Price","Daily_RSI","Weekly_RSI","Monthly_RSI"]
    ws2.append(summary_cols)
    _style_header_row(ws2, len(summary_cols))

    for _, row in df_rsi.iterrows():
        d_rsi = row["Daily_RSI"]; w_rsi = row["Weekly_RSI"]; m_rsi = row["Monthly_RSI"]
        ws2.append([row["Symbol"], "", row["Current_Price"], d_rsi, w_rsi, m_rsi])
        try:
            highlight = (m_rsi is not None and m_rsi > 60 and
                         w_rsi is not None and w_rsi > 60 and
                         d_rsi is not None and 40 <= d_rsi < 45)
        except TypeError:
            highlight = False
        data_row = ws2.max_row
        _style_data_row(ws2, data_row, len(summary_cols), fill=LIGHT_GREEN if highlight else None)
        tv_url = f"https://www.tradingview.com/chart/?symbol=NSE:{row['Symbol']}"
        chart_cell = ws2.cell(row=data_row, column=2)
        chart_cell.hyperlink = tv_url; chart_cell.value = "Chart"
        chart_cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")
        chart_cell.alignment = CENTER_ALIGN

    _set_col_widths(ws2, [16,9,14,11,13,14])
    wb.save(output_path)


def write_json(results, rsi_all, output_path, label):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    payload = {
        "label":       label,
        "generated":   datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"),
        "screened":    results,
        "rsi_summary": rsi_all,
    }
    with open(output_path, "w") as f:
        json.dump(payload, f, indent=2, default=str)


def main():
    print(f"Loading symbols from: {INPUT_CSV}")
    symbols = load_symbol_list(INPUT_CSV)
    print(f"Total symbols: {len(symbols)}\n")

    results = []
    rsi_all = []

    for i, sym in enumerate(symbols, 1):
        print(f"[{i:>3}/{len(symbols)}] {sym} …", end=" ", flush=True)
        rsi_row, screen_row = fetch_and_process(sym)
        if rsi_row:    rsi_all.append(rsi_row)
        if screen_row:
            results.append(screen_row)
            print(f"✅  D:{screen_row['Daily_RSI']}  W:{screen_row['Weekly_RSI']}  M:{screen_row['Monthly_RSI']}")
        elif rsi_row:
            print(f"–   D:{rsi_row['Daily_RSI']}  W:{rsi_row['Weekly_RSI']}  M:{rsi_row['Monthly_RSI']}")
        else:
            print("–  (no data)")
        time.sleep(0.3)

    if not rsi_all:
        print("No data retrieved."); return

    write_excel(results, rsi_all, OUTPUT_XLSX)
    write_json(results, rsi_all, OUTPUT_JSON, "Nifty 50")

    print(f"\n{'='*60}")
    print(f"✅  {len(results)} stock(s) passed all screener filters.")
    print(f"📊  {len(rsi_all)} stock(s) in RSI Summary.")
    print(f"📄  Excel saved: {OUTPUT_XLSX}")
    print(f"📄  JSON  saved: {OUTPUT_JSON}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
